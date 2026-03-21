"""
VELORIX — Modul Export Contabil
=================================
Genereaza exporturi in format CSV, Excel (.xlsx) si PDF
pentru contabil. Contine:
  - Lista devize emise
  - Centralizator TVA colectat
  - Venituri manopera vs piese
  - Situatie incasari pe client
  - Raport lunar rezumat
"""

import os
import csv
from datetime import datetime

# ─── openpyxl ────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ─── reportlab (PDF) ─────────────────────────────────────────
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from database import get_connection, get_tva


# ═══════════════════════════════════════════════════════════════
#  CULORI & STILURI EXCEL
# ═══════════════════════════════════════════════════════════════
NAVY      = "0D1F3C"
BLUE      = "1A73E8"
GREEN     = "10B981"
AMBER     = "F59E0B"
GRAY_LITE = "F0F4F8"
GRAY_MID  = "DCE4EF"
WHITE     = "FFFFFF"

def _hdr_style(ws, row, cols, bg=NAVY, fg=WHITE, bold=True, height=22):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=bold, color=fg, size=10, name="Arial")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color=GRAY_MID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[row].height = height

def _row_style(ws, row, cols, bg=WHITE, bold=False, align_right_cols=None):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=bold, color="1A2535", size=9, name="Arial")
    thin = Side(style="thin", color=GRAY_MID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_right_cols = align_right_cols or []
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(
            horizontal="right" if c in align_right_cols else "center",
            vertical="center"
        )

def _total_style(ws, row, cols, align_right_cols=None):
    fill = PatternFill("solid", fgColor=GRAY_LITE)
    font = Font(bold=True, color=NAVY, size=9, name="Arial")
    thin = Side(style="thin", color=NAVY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_right_cols = align_right_cols or []
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(
            horizontal="right" if c in align_right_cols else "center",
            vertical="center"
        )


# ═══════════════════════════════════════════════════════════════
#  EXTRAGERE DATE DIN DB
# ═══════════════════════════════════════════════════════════════

def _get_data(data_start, data_end):
    """Returneaza toate datele necesare pentru perioada selectata."""
    con = get_connection()
    cur = con.cursor()
    tva_pct = get_tva() / 100

    # ── 1. Lista devize ──
    cur.execute("""
        SELECT d.numar, d.data,
               c.nume AS client,
               v.marca || ' ' || COALESCE(v.model,'') || ' ' || COALESCE(v.nr,'') AS vehicul,
               COALESCE(d.total_manopera, 0),
               COALESCE(d.total_tva, 0),
               COALESCE(d.total_general, 0)
        FROM devize d
        JOIN clienti c ON c.id = d.id_client
        JOIN vehicule v ON v.id = d.id_vehicul
        WHERE d.data BETWEEN ? AND ?
        ORDER BY d.data, d.numar
    """, (data_start, data_end))
    devize = cur.fetchall()

    # ── 2. Centralizator TVA ──
    cur.execute("""
        SELECT
            strftime('%Y-%m', d.data) AS luna,
            SUM(COALESCE(d.total_manopera,0))  AS baza_manopera,
            SUM(COALESCE(dp.total_piese,0))    AS baza_piese,
            SUM(COALESCE(d.total_tva,0))       AS tva_total,
            SUM(COALESCE(d.total_general,0))   AS total_general
        FROM devize d
        LEFT JOIN (
            SELECT id_deviz, SUM(COALESCE(pret_fara_tva*cantitate,0)) AS total_piese
            FROM deviz_piese GROUP BY id_deviz
        ) dp ON dp.id_deviz = d.id
        WHERE d.data BETWEEN ? AND ?
        GROUP BY luna
        ORDER BY luna
    """, (data_start, data_end))
    tva_lunar = cur.fetchall()

    # ── 3. Manopera vs Piese ──
    cur.execute("""
        SELECT d.numar, d.data,
               COALESCE(d.total_manopera, 0) AS manopera,
               COALESCE(SUM(dp.pret_fara_tva * dp.cantitate), 0) AS piese,
               COALESCE(d.total_general, 0)  AS total
        FROM devize d
        LEFT JOIN deviz_piese dp ON dp.id_deviz = d.id
        WHERE d.data BETWEEN ? AND ?
        GROUP BY d.id
        ORDER BY d.data
    """, (data_start, data_end))
    manopera_piese = cur.fetchall()

    # ── 4. Situatie pe client ──
    cur.execute("""
        SELECT c.nume, c.telefon,
               COUNT(d.id)                           AS nr_devize,
               SUM(COALESCE(d.total_manopera,0))     AS total_manopera,
               SUM(COALESCE(d.total_tva,0))          AS total_tva,
               SUM(COALESCE(d.total_general,0))      AS total_general
        FROM devize d
        JOIN clienti c ON c.id = d.id_client
        WHERE d.data BETWEEN ? AND ?
        GROUP BY c.id
        ORDER BY total_general DESC
    """, (data_start, data_end))
    per_client = cur.fetchall()

    # ── 5. Rezumat lunar ──
    cur.execute("""
        SELECT strftime('%Y-%m', d.data)        AS luna,
               COUNT(d.id)                      AS nr_devize,
               SUM(COALESCE(d.total_manopera,0)) AS manopera,
               SUM(COALESCE(d.total_tva,0))     AS tva,
               SUM(COALESCE(d.total_general,0)) AS total
        FROM devize d
        WHERE d.data BETWEEN ? AND ?
        GROUP BY luna
        ORDER BY luna
    """, (data_start, data_end))
    lunar = cur.fetchall()

    # ── Info firma ──
    cur.execute("SELECT nume, cui, adresa, telefon FROM firma LIMIT 1")
    firma = cur.fetchone() or ("—", "—", "—", "—")

    con.close()
    return {
        "devize":         devize,
        "tva_lunar":      tva_lunar,
        "manopera_piese": manopera_piese,
        "per_client":     per_client,
        "lunar":          lunar,
        "firma":          firma,
        "tva_pct":        tva_pct,
        "data_start":     data_start,
        "data_end":       data_end,
    }


# ═══════════════════════════════════════════════════════════════
#  EXPORT CSV
# ═══════════════════════════════════════════════════════════════

def export_csv(data_start, data_end, path_dir):
    """Genereaza un set de fisiere CSV in directorul dat."""
    d = _get_data(data_start, data_end)
    fisiere = []

    def _write(filename, headers, rows):
        fpath = os.path.join(path_dir, filename)
        with open(fpath, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";", quoting=csv.QUOTE_ALL)
            w.writerow(headers)
            for r in rows:
                w.writerow(r)
        fisiere.append(fpath)

    # 1. Devize
    _write("01_devize.csv",
        ["Nr. Deviz", "Data", "Client", "Vehicul",
         "Baza impozabila (RON)", "TVA (RON)", "Total (RON)"],
        [(r[0], r[1], r[2], r[3].strip(),
          f"{r[4]:.2f}", f"{r[5]:.2f}", f"{r[6]:.2f}")
         for r in d["devize"]]
    )

    # 2. TVA lunar
    _write("02_centralizator_tva.csv",
        ["Luna", "Baza manopera", "Baza piese",
         "TVA total (RON)", "Total general (RON)"],
        [(r[0], f"{r[1]:.2f}", f"{r[2]:.2f}",
          f"{r[3]:.2f}", f"{r[4]:.2f}")
         for r in d["tva_lunar"]]
    )

    # 3. Manopera vs piese
    _write("03_manopera_vs_piese.csv",
        ["Nr. Deviz", "Data", "Manopera (RON)",
         "Piese (RON)", "Total (RON)"],
        [(r[0], r[1], f"{r[2]:.2f}", f"{r[3]:.2f}", f"{r[4]:.2f}")
         for r in d["manopera_piese"]]
    )

    # 4. Per client
    _write("04_situatie_clienti.csv",
        ["Client", "Telefon", "Nr. devize",
         "Manopera (RON)", "TVA (RON)", "Total (RON)"],
        [(r[0], r[1] or "—", r[2],
          f"{r[3]:.2f}", f"{r[4]:.2f}", f"{r[5]:.2f}")
         for r in d["per_client"]]
    )

    # 5. Rezumat lunar
    _write("05_rezumat_lunar.csv",
        ["Luna", "Nr. devize", "Manopera (RON)",
         "TVA (RON)", "Total (RON)"],
        [(r[0], r[1], f"{r[2]:.2f}", f"{r[3]:.2f}", f"{r[4]:.2f}")
         for r in d["lunar"]]
    )

    return fisiere


# ═══════════════════════════════════════════════════════════════
#  EXPORT EXCEL
# ═══════════════════════════════════════════════════════════════

def export_excel(data_start, data_end, path):
    """Genereaza un fisier Excel cu 5 foi de calcul."""
    d = _get_data(data_start, data_end)
    wb = Workbook()

    # ── Foaie 1: Devize ───────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Devize"
    _sheet_devize(ws1, d)

    # ── Foaie 2: Centralizator TVA ────────────────────────────
    ws2 = wb.create_sheet("TVA Lunar")
    _sheet_tva(ws2, d)

    # ── Foaie 3: Manopera vs Piese ────────────────────────────
    ws3 = wb.create_sheet("Manopera vs Piese")
    _sheet_manopera_piese(ws3, d)

    # ── Foaie 4: Situatie Clienti ─────────────────────────────
    ws4 = wb.create_sheet("Situatie Clienti")
    _sheet_clienti(ws4, d)

    # ── Foaie 5: Rezumat Lunar ────────────────────────────────
    ws5 = wb.create_sheet("Rezumat Lunar")
    _sheet_rezumat(ws5, d)

    wb.save(path)
    return path


def _write_title_block(ws, firma, perioada, col_span):
    """Scrie antetul comun in fiecare foaie."""
    # Rand 1 — Firma
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=col_span)
    c = ws.cell(row=1, column=1,
                value=f"VELORIX — {firma[0] or 'Service Moto'}  |  CUI: {firma[1] or '—'}")
    c.font = Font(bold=True, color=WHITE, size=11, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Rand 2 — Perioada
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=col_span)
    c2 = ws.cell(row=2, column=1,
                 value=f"Export contabil  |  Perioada: {perioada}  |  "
                       f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    c2.font = Font(italic=True, color="7A8BA0", size=9, name="Arial")
    c2.fill = PatternFill("solid", fgColor=GRAY_LITE)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18


def _sheet_devize(ws, d):
    firma   = d["firma"]
    perioada = f"{d['data_start']} → {d['data_end']}"
    rows_data = d["devize"]
    N = 7

    _write_title_block(ws, firma, perioada, N)

    # Header
    headers = ["Nr. Deviz", "Data", "Client", "Vehicul",
               "Baza impozabila\n(RON)", "TVA\n(RON)", "Total\n(RON)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _hdr_style(ws, 3, range(1, N+1), height=30)

    # Date
    total_baza = total_tva = total_gen = 0.0
    for i, row in enumerate(rows_data, 4):
        vals = [row[0], row[1], row[2], row[3].strip(),
                float(row[4]), float(row[5]), float(row[6])]
        for c, v in enumerate(vals, 1):
            ws.cell(row=i, column=c, value=v)
        bg = GRAY_LITE if i % 2 == 0 else WHITE
        _row_style(ws, i, range(1, N+1), bg=bg, align_right_cols=[5, 6, 7])
        for c in [5, 6, 7]:
            ws.cell(row=i, column=c).number_format = '#,##0.00'
        total_baza += float(row[4])
        total_tva  += float(row[5])
        total_gen  += float(row[6])

    # Total
    tr = len(rows_data) + 4
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.merge_cells(start_row=tr, start_column=1,
                   end_row=tr, end_column=4)
    ws.cell(row=tr, column=5, value=f'=SUM(E4:E{tr-1})')
    ws.cell(row=tr, column=6, value=f'=SUM(F4:F{tr-1})')
    ws.cell(row=tr, column=7, value=f'=SUM(G4:G{tr-1})')
    _total_style(ws, tr, range(1, N+1), align_right_cols=[5, 6, 7])
    for c in [5, 6, 7]:
        ws.cell(row=tr, column=c).number_format = '#,##0.00'

    # Latimi coloane
    for col, w in zip(range(1, N+1), [14, 11, 28, 24, 16, 14, 14]):
        ws.column_dimensions[get_column_letter(col)].width = w


def _sheet_tva(ws, d):
    firma    = d["firma"]
    perioada = f"{d['data_start']} → {d['data_end']}"
    rows_data = d["tva_lunar"]
    N = 5

    _write_title_block(ws, firma, perioada, N)

    headers = ["Luna", "Baza manopera\n(RON)", "Baza piese\n(RON)",
               "TVA total\n(RON)", "Total general\n(RON)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _hdr_style(ws, 3, range(1, N+1), height=30)

    for i, row in enumerate(rows_data, 4):
        vals = [row[0], float(row[1]), float(row[2]),
                float(row[3]), float(row[4])]
        for c, v in enumerate(vals, 1):
            ws.cell(row=i, column=c, value=v)
        bg = GRAY_LITE if i % 2 == 0 else WHITE
        _row_style(ws, i, range(1, N+1), bg=bg, align_right_cols=[2, 3, 4, 5])
        for c in [2, 3, 4, 5]:
            ws.cell(row=i, column=c).number_format = '#,##0.00'

    tr = len(rows_data) + 4
    ws.cell(row=tr, column=1, value="TOTAL")
    for c in range(2, N+1):
        col_letter = get_column_letter(c)
        ws.cell(row=tr, column=c,
                value=f'=SUM({col_letter}4:{col_letter}{tr-1})')
        ws.cell(row=tr, column=c).number_format = '#,##0.00'
    _total_style(ws, tr, range(1, N+1), align_right_cols=[2, 3, 4, 5])

    for col, w in zip(range(1, N+1), [12, 18, 16, 16, 18]):
        ws.column_dimensions[get_column_letter(col)].width = w


def _sheet_manopera_piese(ws, d):
    firma    = d["firma"]
    perioada = f"{d['data_start']} → {d['data_end']}"
    rows_data = d["manopera_piese"]
    N = 5

    _write_title_block(ws, firma, perioada, N)

    headers = ["Nr. Deviz", "Data", "Manopera\n(RON)",
               "Piese\n(RON)", "Total\n(RON)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _hdr_style(ws, 3, range(1, N+1), bg=BLUE, height=30)

    for i, row in enumerate(rows_data, 4):
        vals = [row[0], row[1], float(row[2]),
                float(row[3]), float(row[4])]
        for c, v in enumerate(vals, 1):
            ws.cell(row=i, column=c, value=v)
        bg = GRAY_LITE if i % 2 == 0 else WHITE
        _row_style(ws, i, range(1, N+1), bg=bg, align_right_cols=[3, 4, 5])
        for c in [3, 4, 5]:
            ws.cell(row=i, column=c).number_format = '#,##0.00'

    tr = len(rows_data) + 4
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.merge_cells(start_row=tr, start_column=1,
                   end_row=tr, end_column=2)
    for c in [3, 4, 5]:
        col_letter = get_column_letter(c)
        ws.cell(row=tr, column=c,
                value=f'=SUM({col_letter}4:{col_letter}{tr-1})')
        ws.cell(row=tr, column=c).number_format = '#,##0.00'
    _total_style(ws, tr, range(1, N+1), align_right_cols=[3, 4, 5])

    for col, w in zip(range(1, N+1), [16, 11, 16, 14, 14]):
        ws.column_dimensions[get_column_letter(col)].width = w


def _sheet_clienti(ws, d):
    firma    = d["firma"]
    perioada = f"{d['data_start']} → {d['data_end']}"
    rows_data = d["per_client"]
    N = 6

    _write_title_block(ws, firma, perioada, N)

    headers = ["Client", "Telefon", "Nr. devize",
               "Manopera\n(RON)", "TVA\n(RON)", "Total\n(RON)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _hdr_style(ws, 3, range(1, N+1), bg=GREEN, height=30)

    for i, row in enumerate(rows_data, 4):
        vals = [row[0], row[1] or "—", int(row[2]),
                float(row[3]), float(row[4]), float(row[5])]
        for c, v in enumerate(vals, 1):
            ws.cell(row=i, column=c, value=v)
        bg = GRAY_LITE if i % 2 == 0 else WHITE
        _row_style(ws, i, range(1, N+1), bg=bg, align_right_cols=[3, 4, 5, 6])
        for c in [4, 5, 6]:
            ws.cell(row=i, column=c).number_format = '#,##0.00'

    tr = len(rows_data) + 4
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.merge_cells(start_row=tr, start_column=1,
                   end_row=tr, end_column=2)
    ws.cell(row=tr, column=3, value=f'=SUM(C4:C{tr-1})')
    for c in [4, 5, 6]:
        col_letter = get_column_letter(c)
        ws.cell(row=tr, column=c,
                value=f'=SUM({col_letter}4:{col_letter}{tr-1})')
        ws.cell(row=tr, column=c).number_format = '#,##0.00'
    _total_style(ws, tr, range(1, N+1), align_right_cols=[3, 4, 5, 6])

    for col, w in zip(range(1, N+1), [28, 14, 11, 16, 14, 14]):
        ws.column_dimensions[get_column_letter(col)].width = w


def _sheet_rezumat(ws, d):
    firma    = d["firma"]
    perioada = f"{d['data_start']} → {d['data_end']}"
    rows_data = d["lunar"]
    N = 5

    _write_title_block(ws, firma, perioada, N)

    headers = ["Luna", "Nr. devize", "Manopera\n(RON)",
               "TVA\n(RON)", "Total\n(RON)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    _hdr_style(ws, 3, range(1, N+1), bg=AMBER, fg="1A2535", height=30)

    for i, row in enumerate(rows_data, 4):
        vals = [row[0], int(row[1]), float(row[2]),
                float(row[3]), float(row[4])]
        for c, v in enumerate(vals, 1):
            ws.cell(row=i, column=c, value=v)
        bg = GRAY_LITE if i % 2 == 0 else WHITE
        _row_style(ws, i, range(1, N+1), bg=bg, align_right_cols=[2, 3, 4, 5])
        for c in [3, 4, 5]:
            ws.cell(row=i, column=c).number_format = '#,##0.00'

    tr = len(rows_data) + 4
    ws.cell(row=tr, column=1, value="TOTAL")
    ws.cell(row=tr, column=2, value=f'=SUM(B4:B{tr-1})')
    for c in [3, 4, 5]:
        col_letter = get_column_letter(c)
        ws.cell(row=tr, column=c,
                value=f'=SUM({col_letter}4:{col_letter}{tr-1})')
        ws.cell(row=tr, column=c).number_format = '#,##0.00'
    _total_style(ws, tr, range(1, N+1), align_right_cols=[2, 3, 4, 5])

    for col, w in zip(range(1, N+1), [12, 12, 18, 14, 16]):
        ws.column_dimensions[get_column_letter(col)].width = w


# ═══════════════════════════════════════════════════════════════
#  EXPORT PDF
# ═══════════════════════════════════════════════════════════════

def export_pdf(data_start, data_end, path):
    """Genereaza un PDF cu toate rapoartele contabile."""
    d   = _get_data(data_start, data_end)
    doc = SimpleDocTemplate(
        path,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
        title="VELORIX — Export Contabil"
    )

    styles = getSampleStyleSheet()
    NAVY_C  = colors.HexColor("#0D1F3C")
    BLUE_C  = colors.HexColor("#1A73E8")
    GREEN_C = colors.HexColor("#10B981")
    AMBER_C = colors.HexColor("#F59E0B")
    GRAY_C  = colors.HexColor("#F0F4F8")

    def style_h1(text):
        return Paragraph(
            f'<font color="#0D1F3C" size="14"><b>{text}</b></font>', styles["Normal"]
        )

    def style_h2(text, color="#1A73E8"):
        return Paragraph(
            f'<font color="{color}" size="11"><b>{text}</b></font>', styles["Normal"]
        )

    def style_meta(text):
        return Paragraph(
            f'<font color="#7A8BA0" size="8">{text}</font>', styles["Normal"]
        )

    def make_table(headers, rows, col_widths, hdr_color=NAVY_C):
        data = [headers] + rows
        t = Table(data, colWidths=col_widths, repeatRows=1)
        style = TableStyle([
            # Header
            ("BACKGROUND",  (0, 0), (-1, 0), hdr_color),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0), 8),
            ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
            ("VALIGN",      (0, 0), (-1, 0), "MIDDLE"),
            ("ROWBACKGROUND",(0, 0), (-1, 0), hdr_color),
            ("TOPPADDING",  (0, 0), (-1, 0), 6),
            ("BOTTOMPADDING",(0, 0), (-1, 0), 6),
            # Body
            ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",    (0, 1), (-1, -1), 7.5),
            ("ALIGN",       (0, 1), (-1, -1), "CENTER"),
            ("VALIGN",      (0, 1), (-1, -1), "MIDDLE"),
            ("TOPPADDING",  (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING",(0, 1), (-1, -1), 4),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F0F4F8")]),
            # Grid
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#DCE4EF")),
            ("LINEBELOW",   (0, 0), (-1, 0), 1, hdr_color),
        ])
        t.setStyle(style)
        return t

    # ── Antet ──
    firma    = d["firma"]
    perioada = f"{data_start} → {data_end}"
    story = [
        style_h1(f"VELORIX — Export Contabil"),
        Spacer(1, 4),
        style_meta(
            f"{firma[0] or 'Service Moto'}  |  CUI: {firma[1] or '—'}  |  "
            f"Perioada: {perioada}  |  "
            f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ),
        HRFlowable(width="100%", thickness=1,
                   color=colors.HexColor("#1A73E8"), spaceAfter=12),
    ]

    W = landscape(A4)[0] - 3*cm   # latime utila

    # ── 1. Devize ──
    story.append(style_h2("1. Lista Devize Emise", "#0D1F3C"))
    story.append(Spacer(1, 6))
    dev_rows = [
        [r[0], r[1], r[2][:22] if r[2] else "—",
         r[3].strip()[:20] if r[3] else "—",
         f"{float(r[4]):.2f}", f"{float(r[5]):.2f}", f"{float(r[6]):.2f}"]
        for r in d["devize"]
    ]
    if dev_rows:
        story.append(make_table(
            ["Nr. Deviz", "Data", "Client", "Vehicul",
             "Baza (RON)", "TVA (RON)", "Total (RON)"],
            dev_rows,
            [W*0.13, W*0.09, W*0.22, W*0.20,
             W*0.13, W*0.11, W*0.12]
        ))
    else:
        story.append(style_meta("— Nu exista devize in aceasta perioada —"))
    story.append(Spacer(1, 14))

    # ── 2. TVA Lunar ──
    story.append(style_h2("2. Centralizator TVA Lunar", "#1A73E8"))
    story.append(Spacer(1, 6))
    tva_rows = [
        [r[0], f"{float(r[1]):.2f}", f"{float(r[2]):.2f}",
         f"{float(r[3]):.2f}", f"{float(r[4]):.2f}"]
        for r in d["tva_lunar"]
    ]
    if tva_rows:
        story.append(make_table(
            ["Luna", "Baza manopera (RON)", "Baza piese (RON)",
             "TVA (RON)", "Total (RON)"],
            tva_rows,
            [W*0.12, W*0.22, W*0.20, W*0.22, W*0.24],
            hdr_color=BLUE_C
        ))
    else:
        story.append(style_meta("— Nu exista date TVA in aceasta perioada —"))
    story.append(Spacer(1, 14))

    # ── 3. Manopera vs Piese ──
    story.append(style_h2("3. Venituri Manopera vs Piese", "#0D1F3C"))
    story.append(Spacer(1, 6))
    mp_rows = [
        [r[0], r[1], f"{float(r[2]):.2f}",
         f"{float(r[3]):.2f}", f"{float(r[4]):.2f}"]
        for r in d["manopera_piese"]
    ]
    if mp_rows:
        story.append(make_table(
            ["Nr. Deviz", "Data", "Manopera (RON)",
             "Piese (RON)", "Total (RON)"],
            mp_rows,
            [W*0.16, W*0.12, W*0.24, W*0.22, W*0.26]
        ))
    else:
        story.append(style_meta("— Nu exista date in aceasta perioada —"))
    story.append(Spacer(1, 14))

    # ── 4. Situatie Clienti ──
    story.append(style_h2("4. Situatie Incasari pe Client", "#10B981"))
    story.append(Spacer(1, 6))
    cl_rows = [
        [r[0][:25] if r[0] else "—", r[1] or "—", str(r[2]),
         f"{float(r[3]):.2f}", f"{float(r[4]):.2f}", f"{float(r[5]):.2f}"]
        for r in d["per_client"]
    ]
    if cl_rows:
        story.append(make_table(
            ["Client", "Telefon", "Nr. devize",
             "Manopera (RON)", "TVA (RON)", "Total (RON)"],
            cl_rows,
            [W*0.24, W*0.14, W*0.10,
             W*0.18, W*0.16, W*0.18],
            hdr_color=GREEN_C
        ))
    else:
        story.append(style_meta("— Nu exista clienti in aceasta perioada —"))
    story.append(Spacer(1, 14))

    # ── 5. Rezumat Lunar ──
    story.append(style_h2("5. Raport Lunar Rezumat", "#F59E0B"))
    story.append(Spacer(1, 6))
    rez_rows = [
        [r[0], str(r[1]), f"{float(r[2]):.2f}",
         f"{float(r[3]):.2f}", f"{float(r[4]):.2f}"]
        for r in d["lunar"]
    ]

    # Total general
    total_dev  = sum(r[1] for r in d["lunar"])
    total_man  = sum(float(r[2]) for r in d["lunar"])
    total_tva2 = sum(float(r[3]) for r in d["lunar"])
    total_gen  = sum(float(r[4]) for r in d["lunar"])
    rez_rows.append([
        "TOTAL", str(total_dev),
        f"{total_man:.2f}", f"{total_tva2:.2f}", f"{total_gen:.2f}"
    ])

    if rez_rows:
        t = make_table(
            ["Luna", "Nr. devize", "Manopera (RON)",
             "TVA (RON)", "Total (RON)"],
            rez_rows,
            [W*0.14, W*0.14, W*0.24, W*0.22, W*0.26],
            hdr_color=AMBER_C
        )
        # Stilizare rand TOTAL
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, len(rez_rows)), (-1, len(rez_rows)),
             colors.HexColor("#F0F4F8")),
            ("FONTNAME",    (0, len(rez_rows)), (-1, len(rez_rows)),
             "Helvetica-Bold"),
            ("TEXTCOLOR",   (0, len(rez_rows)), (-1, len(rez_rows)),
             colors.HexColor("#0D1F3C")),
            ("LINEABOVE",   (0, len(rez_rows)), (-1, len(rez_rows)),
             1, colors.HexColor("#0D1F3C")),
        ]))
        story.append(t)

    doc.build(story)
    return path